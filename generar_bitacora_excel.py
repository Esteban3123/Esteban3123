#!/usr/bin/env python3
"""Genera Bitácora GFPI-F-147 en Excel (.xlsx) lista para entregar."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
from openpyxl.utils import get_column_letter


GREEN = PatternFill("solid", fgColor="1F4E3D")
LIGHT_GREEN = PatternFill("solid", fgColor="D9EAD3")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
GRAY = PatternFill("solid", fgColor="F3F3F3")
WHITE = PatternFill("solid", fgColor="FFFFFF")

THIN = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)


def style_cell(cell, *, bold=False, size=9, fill=None, wrap=True, center=False, color="000000"):
    cell.font = Font(name="Arial", size=size, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=wrap,
    )
    cell.border = THIN
    if fill:
        cell.fill = fill


def merge_write(ws, r1, c1, r2, c2, value, **kwargs):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    cell = ws.cell(r1, c1, value)
    style_cell(cell, **kwargs)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = THIN
            if kwargs.get("fill"):
                ws.cell(r, c).fill = kwargs["fill"]
    return cell


def generar_excel(datos, actividades, salida):
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bitacora {datos['bitacora_n']}"

    # Anchos
    widths = [28, 28, 18, 18, 22, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Encabezado ---
    merge_write(ws, 1, 1, 1, 2, "Código:\nGFPI-F-147", bold=True, size=9, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 1, 3, 1, 4, "Versión: 05", bold=True, size=9, fill=LIGHT_GREEN, center=True)
    merge_write(
        ws, 1, 5, 1, 6,
        "PROCESO\nGESTIÓN DE FORMACIÓN PROFESIONAL INTEGRAL",
        bold=True, size=8, fill=LIGHT_GREEN, center=True,
    )
    ws.row_dimensions[1].height = 36

    merge_write(
        ws, 2, 1, 2, 6,
        "FORMATO BITÁCORA DE SEGUIMIENTO ETAPA PRODUCTIVA",
        bold=True, size=11, fill=GREEN, center=True, color="FFFFFF",
    )
    ws.row_dimensions[2].height = 28

    merge_write(ws, 3, 1, 3, 6, "CLASIFICACIÓN DE LA INFORMACIÓN: Pública Clasificada  X", bold=True, size=9, fill=GRAY, center=True)

    # --- Bitácora / periodo ---
    r = 5
    merge_write(ws, r, 1, r, 1, "Bitácora N°", bold=True, size=9, fill=LIGHT_GREEN, center=True)
    merge_write(ws, r, 2, r, 2, datos["bitacora_n"], bold=True, size=14, fill=YELLOW, center=True)
    merge_write(ws, r, 3, r, 3, "Desde", bold=True, size=9, fill=LIGHT_GREEN, center=True)
    merge_write(ws, r, 4, r, 4, datos["periodo_desde"], bold=True, size=10, fill=YELLOW, center=True)
    merge_write(ws, r, 5, r, 5, "Hasta", bold=True, size=9, fill=LIGHT_GREEN, center=True)
    merge_write(ws, r, 6, r, 6, datos["periodo_hasta"], bold=True, size=10, fill=YELLOW, center=True)

    # --- Datos aprendiz ---
    r = 7
    merge_write(ws, r, 1, r, 6, "DATOS DEL APRENDIZ", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")

    labels = [
        (8, [("Nombre completo", datos["nombre"]), ("Tipo doc.", datos["tipo_doc"]), ("N° identificación", datos["num_doc"])]),
        (9, [("Teléfono", datos["telefono"]), ("Correo institucional", datos["correo_inst"]), ("Correo personal", datos["correo_pers"])]),
        (10, [("Dirección", datos["direccion"]), ("N° grupo / ficha", datos["grupo"]), ("Modalidad formación", datos["modalidad_formacion"])]),
        (11, [("Programa", datos["programa"]), ("Modalidad etapa productiva", datos["modalidad_etapa"]), ("Exterior / País", f"{datos['exterior']} / {datos['pais']}")]),
    ]
    for row, pairs in labels:
        c = 1
        for lab, val in pairs:
            merge_write(ws, row, c, row, c, lab, bold=True, size=8, fill=LIGHT_GREEN, center=True)
            merge_write(ws, row, c + 1, row, c + 1, val, size=9, fill=WHITE, center=True)
            c += 2
        ws.row_dimensions[row].height = 22

    # --- Empresa ---
    r = 13
    merge_write(ws, r, 1, r, 6, "DATOS DE LA ENTIDAD CO-FORMADORA", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")
    merge_write(ws, 14, 1, 14, 1, "Empresa", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 14, 2, 14, 2, datos["empresa"], size=9, center=True)
    merge_write(ws, 14, 3, 14, 3, "NIT", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 14, 4, 14, 4, datos["nit"], size=9, center=True)
    merge_write(ws, 14, 5, 14, 5, "Dirección", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 14, 6, 14, 6, datos["dir_empresa"], size=9, center=True)

    merge_write(ws, 15, 1, 15, 1, "Jefe / Supervisor", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 15, 2, 15, 2, datos["jefe_nombre"], size=9, center=True)
    merge_write(ws, 15, 3, 15, 3, "Cargo", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 15, 4, 15, 4, datos["jefe_cargo"], size=9, center=True)
    merge_write(ws, 15, 5, 15, 5, "Tel / Correo", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 15, 6, 15, 6, f"{datos['jefe_tel']} / {datos['jefe_correo']}", size=8, center=True)

    # --- Instructor ---
    merge_write(ws, 17, 1, 17, 6, "DATOS DEL INSTRUCTOR DE SEGUIMIENTO", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")
    merge_write(ws, 18, 1, 18, 1, "Nombre", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 18, 2, 18, 3, datos["instructor_nombre"], size=9, center=True)
    merge_write(ws, 18, 4, 18, 4, "Correo", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 18, 5, 18, 6, datos["instructor_correo"], size=9, center=True)

    # --- Alternativa ---
    merge_write(ws, 20, 1, 20, 6, "ALTERNATIVA DE ETAPA PRODUCTIVA", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")
    alternativas = [
        "Contrato de aprendizaje",
        "Monitoria",
        "Proyecto productivo",
        "Contrato de vínculo formativo",
        "Vínculo laboral",
    ]
    merge_write(ws, 21, 1, 21, 4, "Alternativa", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, 21, 5, 21, 6, 'Marque con "X"', bold=True, size=8, fill=LIGHT_GREEN, center=True)
    for i, alt in enumerate(alternativas):
        row = 22 + i
        merge_write(ws, row, 1, row, 4, alt, size=9)
        marca = "X" if alt.lower() == datos["alternativa"].lower() else ""
        merge_write(ws, row, 5, row, 6, marca, bold=True, size=12, fill=YELLOW if marca else WHITE, center=True)

    # --- Actividades ---
    start = 28
    merge_write(ws, start, 1, start, 6, "DESCRIPCIÓN DE LAS ACTIVIDADES REALIZADAS", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")

    headers = [
        "Descripción de la actividad",
        "Competencias del programa aplicadas",
        "Fecha inicio\n(dd/mm/aa)",
        "Fecha fin\n(dd/mm/aa)",
        "Evidencia de cumplimiento",
        "Observaciones / inasistencias / dificultades",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(start + 1, c, h)
        style_cell(cell, bold=True, size=8, fill=LIGHT_GREEN, center=True)
    ws.row_dimensions[start + 1].height = 40

    for i, act in enumerate(actividades):
        row = start + 2 + i
        vals = [
            act.get("descripcion", ""),
            act.get("competencias", ""),
            act.get("fecha_inicio", ""),
            act.get("fecha_fin", ""),
            act.get("evidencia", ""),
            act.get("observaciones", ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row, c, v)
            style_cell(cell, size=8, fill=WHITE, center=(c in (3, 4)))
        ws.row_dimensions[row].height = 70

    # --- ARL ---
    arl_row = start + 2 + len(actividades) + 2
    merge_write(ws, arl_row, 1, arl_row, 6, "INFORMACIÓN AFILIACIÓN A LA ARL (Decreto 055 de 2015)", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")
    arl_headers = [
        "¿Afiliado a la ARL?",
        "Nivel de riesgo",
        "¿El nivel corresponde a las actividades? (SI/NO)",
        "¿Cuenta con EPP? (SI/NO/NA)",
    ]
    # Use 4 cells spanning 6 cols: 1-1, 2-2, 3-4, 5-6
    spans = [(1, 1), (2, 2), (3, 4), (5, 6)]
    vals = [datos["arl_afiliado"], datos["arl_nivel"], datos["arl_corresponde"], datos["arl_epp"]]
    for (c1, c2), h, v in zip(spans, arl_headers, vals):
        merge_write(ws, arl_row + 1, c1, arl_row + 1, c2, h, bold=True, size=8, fill=LIGHT_GREEN, center=True)
        merge_write(ws, arl_row + 2, c1, arl_row + 2, c2, v, bold=True, size=11, fill=YELLOW, center=True)
    ws.row_dimensions[arl_row + 1].height = 32

    # --- Firmas ---
    firmas = arl_row + 4
    merge_write(ws, firmas, 1, firmas, 6, "FIRMAS", bold=True, size=10, fill=GREEN, center=True, color="FFFFFF")
    merge_write(ws, firmas + 1, 1, firmas + 1, 2, "Firma del aprendiz", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, firmas + 1, 3, firmas + 1, 4, "Fecha entrega bitácora", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, firmas + 1, 5, firmas + 1, 6, "Firma instructor / ente co-formador", bold=True, size=8, fill=LIGHT_GREEN, center=True)
    merge_write(ws, firmas + 2, 1, firmas + 2, 2, "\n\n_______________________\nEsteban Andres Perdomo Rojas", size=8, center=True)
    merge_write(ws, firmas + 2, 3, firmas + 2, 4, datos["fecha_entrega"], bold=True, size=12, fill=YELLOW, center=True)
    merge_write(ws, firmas + 2, 5, firmas + 2, 6, "\n\n_______________________\nInstructor / Jefe inmediato", size=8, center=True)
    ws.row_dimensions[firmas + 2].height = 60

    nota = firmas + 4
    merge_write(
        ws, nota, 1, nota, 6,
        "Nota: Con el diligenciamiento de este formato autorizo al SENA para la recolección y tratamiento "
        "de mis datos personales, conforme a la política GOR-POL-006.",
        size=7, fill=GRAY,
    )
    ws.row_dimensions[nota].height = 30

    anexo = nota + 1
    merge_write(
        ws, anexo, 1, anexo, 6,
        "Anexo (opcional): evidencia fotográfica de las actividades. No aplicar documentos sensibles de la empresa.",
        size=8, bold=True,
    )

    wb.save(salida)
    print(f"Generado: {salida}")


if __name__ == "__main__":
    from datos_bitacora import DATOS, ACTIVIDADES

    n = DATOS.get("bitacora_n", "X")
    generar_excel(
        DATOS,
        ACTIVIDADES,
        f"/workspace/Bitacora{n}_Esteban_Andres_Perdomo_Rojas.xlsx",
    )
